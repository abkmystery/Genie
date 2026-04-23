from __future__ import annotations

import csv
import mimetypes
from pathlib import Path
from uuid import uuid4

from docx import Document
from openpyxl import load_workbook
from pypdf import PdfReader

from app.models.contracts import SourceChunk
from app.providers.interfaces import OCRProvider, SourceParser


def _chunk_text(text: str, locator_prefix: str, metadata: dict[str, object], chunk_size: int = 900) -> list[SourceChunk]:
    clean = "\n".join(part.strip() for part in text.splitlines() if part.strip())
    if not clean:
        return []
    chunks: list[SourceChunk] = []
    start = 0
    ordinal = 0
    while start < len(clean):
        chunk_text = clean[start : start + chunk_size]
        chunks.append(
            SourceChunk(
                id=str(uuid4()),
                source_id="",
                ordinal=ordinal,
                text=chunk_text,
                locator=f"{locator_prefix}{ordinal + 1}",
                metadata=dict(metadata),
            )
        )
        start += chunk_size
        ordinal += 1
    return chunks


class TextSourceParser(SourceParser):
    extensions = {".txt", ".md"}

    def supports(self, path: Path) -> bool:
        return path.suffix.lower() in self.extensions

    def parse(self, path: Path) -> list[SourceChunk]:
        return _chunk_text(path.read_text(encoding="utf-8", errors="ignore"), "section=", {"kind": path.suffix.lower()})


class CsvSourceParser(SourceParser):
    def supports(self, path: Path) -> bool:
        return path.suffix.lower() == ".csv"

    def parse(self, path: Path) -> list[SourceChunk]:
        chunks: list[SourceChunk] = []
        with path.open("r", encoding="utf-8", errors="ignore", newline="") as handle:
            reader = csv.reader(handle)
            headers = next(reader, [])
            for idx, row in enumerate(reader, start=1):
                text = ", ".join(f"{header}={value}" for header, value in zip(headers, row, strict=False))
                if not text:
                    continue
                chunks.append(
                    SourceChunk(
                        id=str(uuid4()),
                        source_id="",
                        ordinal=idx - 1,
                        text=text,
                        locator=f"row-range={idx}-{idx}",
                        metadata={"headers": headers},
                    )
                )
        return chunks


class XlsxSourceParser(SourceParser):
    def supports(self, path: Path) -> bool:
        return path.suffix.lower() == ".xlsx"

    def parse(self, path: Path) -> list[SourceChunk]:
        workbook = load_workbook(path, read_only=True, data_only=True)
        chunks: list[SourceChunk] = []
        ordinal = 0
        for sheet in workbook.worksheets:
            for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                values = [str(value) for value in row if value is not None]
                if not values:
                    continue
                chunks.append(
                    SourceChunk(
                        id=str(uuid4()),
                        source_id="",
                        ordinal=ordinal,
                        text=" | ".join(values),
                        locator=f"sheet={sheet.title} row-range={row_index}-{row_index}",
                        metadata={"sheet": sheet.title, "row_index": row_index},
                    )
                )
                ordinal += 1
        return chunks


class PdfSourceParser(SourceParser):
    def supports(self, path: Path) -> bool:
        return path.suffix.lower() == ".pdf"

    def parse(self, path: Path) -> list[SourceChunk]:
        reader = PdfReader(str(path))
        chunks: list[SourceChunk] = []
        ordinal = 0
        for page_index, page in enumerate(reader.pages, start=1):
            text = page.extract_text() or ""
            for chunk in _chunk_text(text, f"p.{page_index}-chunk=", {"page": page_index}):
                chunk.ordinal = ordinal
                ordinal += 1
                chunks.append(chunk)
        return chunks


class DocxSourceParser(SourceParser):
    def supports(self, path: Path) -> bool:
        return path.suffix.lower() == ".docx"

    def parse(self, path: Path) -> list[SourceChunk]:
        document = Document(path)
        text = "\n".join(paragraph.text for paragraph in document.paragraphs if paragraph.text.strip())
        return _chunk_text(text, "section=", {"kind": "docx"})


class ImageSourceParser(SourceParser):
    extensions = {".png", ".jpg", ".jpeg", ".webp"}

    def __init__(self, ocr_provider: OCRProvider) -> None:
        self.ocr_provider = ocr_provider

    def supports(self, path: Path) -> bool:
        return path.suffix.lower() in self.extensions

    def parse(self, path: Path) -> list[SourceChunk]:
        extracted = self.ocr_provider.extract_text(path)
        description = extracted or f"Image source {path.name} with no OCR available."
        return [
            SourceChunk(
                id=str(uuid4()),
                source_id="",
                ordinal=0,
                text=description,
                locator=None,
                metadata={"mime_type": mimetypes.guess_type(path.name)[0] or "image/*"},
            )
        ]


class ParserRegistry:
    def __init__(self, parsers: list[SourceParser]) -> None:
        self.parsers = parsers

    def get_parser(self, path: Path) -> SourceParser:
        for parser in self.parsers:
            if parser.supports(path):
                return parser
        raise ValueError(f"Unsupported file type: {path.suffix}")

